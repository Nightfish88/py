import openpyxl
import os
import requests
from urllib.parse import urlparse
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def download_favicon(url, save_directory):
    # Extract the domain from the URL
    domain = urlparse(url).netloc

    # Construct the Google API link for the favicon
    api_url = f"https://www.google.com/s2/favicons?domain={url}&sz=128"

    # Send a GET request to the Google API
    response = requests.get(api_url)

    # Check if the request was successful
    if response.status_code == 200:
        # Save the favicon as an image file
        favicon_filename = f'{domain}_favicon.ico'
        favicon_filepath = os.path.join(save_directory, favicon_filename)
        with open(favicon_filepath, 'wb') as file:
            file.write(response.content)
        print(f"Favicon for {url} downloaded successfully!")
        return favicon_filepath
    else:
        print(f"Failed to download favicon for {url}.")
        return None


# Load the workbook
workbook = openpyxl.load_workbook('Claim & Restake.xlsx')

# Select the worksheet
worksheet = workbook['Sheet1']

# Extract the links from column C, skipping the first row
links = []
for cell in worksheet['C'][1:]:
    if cell.hyperlink:
        link = cell.hyperlink.target
    else:
        link = cell.value
    links.append(link)

# Create a directory to store the favicon files
os.makedirs('favicons', exist_ok=True)

# Fetch the favicons and save them as image files
for i, link in enumerate(links):
    if link:
        try:
            favicon_filepath = download_favicon(link.strip(), 'favicons')
            if favicon_filepath:
                # Insert the favicon as a picture into the next column (column D)
                img = XLImage(favicon_filepath)
                img.width = 16
                img.height = 16
                worksheet.column_dimensions[get_column_letter(4)].width = 20
                worksheet.row_dimensions[i + 2].height = 20
                img.anchor = f'D{i + 2}'
                cell = worksheet.cell(row=i + 2, column=4)
                # cell.value = "=REPT("" "",(20-LEN(C{0}))&C{0})".format(i+2)  # Insert empty spaces to center the image
                worksheet.add_image(img)
        except Exception as e:
            print(f"An error occurred while processing favicon for {link}: {e}")

# Save the modified workbook
workbook.save('Claim & Restake.xlsx')
